from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
from urp_trigger import send_dashboard_command
from robot_executor import robot_executor
from fastapi import Request
from pydantic import BaseModel
from typing import List
from openpyxl import Workbook, load_workbook
import os
import time
import re
from datetime import datetime
from fastapi import Body

# Set to False when not connected to the robot (for dev/testing)
ROBOT_CONNECTED = True
if not ROBOT_CONNECTED:
    print("‚ö†Ô∏è Robot is in SIMULATED mode ‚Äî no real commands will be sent.")

# Global dependency management
class DependencyManager:
    def __init__(self):
        self.dependencies = {}
        self.robot_task_dependency = []  # Current robot task dependencies
        self.human_task_dependency = []  # Current human task dependencies
        self.robot_assigned_tasks = []   # Tasks assigned to robot
        self.human_assigned_tasks = []   # Tasks assigned to human
        self.all_finished_tasks = []     # All finished tasks
        self.load_dependencies()
    
    def load_dependencies(self):
        """Load all task dependencies from tasks.xlsx"""
        try:
            df = pd.read_excel("tasks.xlsx")
            for _, row in df.iterrows():
                task_name = str(row["TaskName"]).strip()
                robot_code = str(row["RobotCode"]).strip()
                dep = str(row.get("Dependency", "")).strip()
                
                if dep and dep.lower() != "nan":
                    self.dependencies[task_name] = [d.strip() for d in dep.split(",")]
                else:
                    self.dependencies[task_name] = []
                    
            print("üìé Loaded all task dependencies:", self.dependencies)
        except Exception as e:
            print(f"‚ùå Failed to load dependencies: {e}")
            self.dependencies = {}
    
    def update_assigned_tasks(self, tasks):
        """Update robot and human assigned task lists after Start button"""
        self.robot_assigned_tasks = []
        self.human_assigned_tasks = []
        
        for task in tasks:
            if task.get("assignedTo") == "Robot":
                self.robot_assigned_tasks.append(task["name"])
            elif task.get("assignedTo") == "Human":
                self.human_assigned_tasks.append(task["name"])
        
        print(f"üìã Updated assigned tasks:")
        print(f"   Robot assigned tasks: {self.robot_assigned_tasks}")
        print(f"   Human assigned tasks: {self.human_assigned_tasks}")
    
    def set_human_task_dependencies(self, task_name):
        """Set dependencies for current human task"""
        deps = self.dependencies.get(task_name, [])
        self.human_task_dependency = deps
        print(f"üë§ Set human task dependencies for '{task_name}': {self.human_task_dependency}")
    
    def set_robot_task_dependencies(self, task_name):
        """Set dependencies for current robot task"""
        deps = self.dependencies.get(task_name, [])
        self.robot_task_dependency = deps
        print(f"ü§ñ Set robot task dependencies for '{task_name}': {self.robot_task_dependency}")
    
    def check_human_dependencies(self, task_name):
        """Check if human task dependencies are met"""
        deps = self.dependencies.get(task_name, [])
        unmet_deps = []
        
        for dep in deps:
            if dep not in self.all_finished_tasks:
                unmet_deps.append(dep)
                print(f"‚ùå Missing humanTask Dependency: '{dep}' for task '{task_name}'")
        
        if unmet_deps:
            return False, f"Please wait till robot execute these tasks: {', '.join(unmet_deps)}"
        else:
            return True, ""
    
    def check_robot_dependencies(self, task_name):
        """Check if robot task dependencies are met"""
        deps = self.dependencies.get(task_name, [])
        unmet_deps = []
        
        for dep in deps:
            if dep not in self.all_finished_tasks:
                unmet_deps.append(dep)
                print(f"ü§ñ Missing robotTaskDependency: '{dep}' for task '{task_name}'")
        
        if unmet_deps:
            return False, f"I am waiting for my task dependencies, I will start after you execute: {', '.join(unmet_deps)}"
        else:
            return True, ""
    
    def add_to_all_finished_tasks(self, task_name):
        """Add task to all finished tasks list"""
        if task_name not in self.all_finished_tasks:
            self.all_finished_tasks.append(task_name)
            print(f"‚úÖ Added '{task_name}' to all finished tasks")
        else:
            print(f"‚ö†Ô∏è Task '{task_name}' already in all finished tasks")
    
    def remove_from_human_assigned_tasks(self, task_name):
        """Remove task from human assigned tasks list"""
        if task_name in self.human_assigned_tasks:
            self.human_assigned_tasks.remove(task_name)
            print(f"‚úÖ Removed '{task_name}' from human assigned tasks")
        else:
            print(f"‚ö†Ô∏è Task '{task_name}' not found in human assigned tasks")
    
    def remove_from_robot_assigned_tasks(self, task_name):
        """Remove task from robot assigned tasks list"""
        if task_name in self.robot_assigned_tasks:
            self.robot_assigned_tasks.remove(task_name)
            print(f"‚úÖ Removed '{task_name}' from robot assigned tasks")
        else:
            print(f"‚ö†Ô∏è Task '{task_name}' not found in robot assigned tasks")
    
    def get_current_human_task(self):
        """Get the first task from human assigned tasks"""
        if self.human_assigned_tasks:
            return self.human_assigned_tasks[0]
        return None
    
    def get_current_robot_task(self):
        """Get the first task from robot assigned tasks"""
        if self.robot_assigned_tasks:
            return self.robot_assigned_tasks[0]
        return None
    
    def reset(self):
        """Reset all lists"""
        self.robot_task_dependency = []
        self.human_task_dependency = []
        self.robot_assigned_tasks = []
        self.human_assigned_tasks = []
        self.all_finished_tasks = []
        print("üîÑ Dependency manager reset - cleared all lists")

# Initialize dependency manager
dependency_manager = DependencyManager()

app = FastAPI()

# Allow frontend React access
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
async def root():
    return {"message": "Robot backend is running!"}

@app.get("/tasks")
def get_tasks(optimizationMode: str = Query("no-optimization")):
    df = pd.read_excel("tasks.xlsx")
    task_list = []
    for _, row in df.iterrows():
        task = {
            "TaskID": int(row["TaskID"]),
            "TaskName": row["TaskName"],
            "Instruction": row["Instruction"],
            "ImageName": row["ImageName"],
            "Time_Human": int(row["Time_Human"]),
            "Time_Robot": int(row["Time_Robot"]),
            "RobotCode": row["RobotCode"],
            "fixedToHuman": str(row["RobotCode"]).strip().lower() == "cannot" if "RobotCode" in df.columns else False
        }
        task_list.append(task)

    # Note: optimize_tasks function is not implemented
    tasks = [{**task, "assignedTo": "Unassigned", "sliderValue": 5} for task in task_list]

    formatted_tasks = []
    for task in tasks:
        formatted_tasks.append({
            "id": str(task["TaskID"]),
            "name": task["TaskName"],
            "description": task["Instruction"],
            "image": f"/{task['ImageName']}",
            "time_human": task["Time_Human"],
            "time_robot": task["Time_Robot"],
            "assignedTo": task["assignedTo"],
            "sliderValue": task["sliderValue"],
            "RobotCode": task["RobotCode"],
            "fixedToHuman": str(task["RobotCode"]).strip().lower() == "cannot"

        })

    return formatted_tasks

@app.get("/robot/message")
def get_robot_message():
    message = robot_executor.robot_message
    execution_message = robot_executor.execution_message
    print(f"üîç Robot message endpoint called, returning: message='{message}', executionMessage='{execution_message}'")
    return {"message": message, "executionMessage": execution_message}



@app.post("/start-execution")
def start_execution(tasks: List[dict]):
    """Start execution - update assigned task lists"""
    dependency_manager.update_assigned_tasks(tasks)
    
    # Set dependencies for current tasks
    current_human_task = dependency_manager.get_current_human_task()
    current_robot_task = dependency_manager.get_current_robot_task()
    
    if current_human_task:
        dependency_manager.set_human_task_dependencies(current_human_task)
    
    if current_robot_task:
        dependency_manager.set_robot_task_dependencies(current_robot_task)
    
    return {
        "status": "success",
        "human_assigned": dependency_manager.human_assigned_tasks,
        "robot_assigned": dependency_manager.robot_assigned_tasks,
        "current_human_task": current_human_task,
        "current_robot_task": current_robot_task
    }

@app.post("/complete-human-task")
def complete_human_task(task_name: str):
    """Complete a human task"""
    # Add to all finished tasks
    dependency_manager.add_to_all_finished_tasks(task_name)
    
    # Remove from human assigned tasks
    dependency_manager.remove_from_human_assigned_tasks(task_name)
    
    # Set dependencies for next human task
    next_human_task = dependency_manager.get_current_human_task()
    if next_human_task:
        dependency_manager.set_human_task_dependencies(next_human_task)
    
    return {
        "status": "success", 
        "task": task_name,
        "next_human_task": next_human_task,
        "human_assigned": dependency_manager.human_assigned_tasks
    }

@app.post("/complete-robot-task")
def complete_robot_task(task_name: str):
    """Complete a robot task"""
    # Add to all finished tasks
    dependency_manager.add_to_all_finished_tasks(task_name)
    
    # Remove from robot assigned tasks
    dependency_manager.remove_from_robot_assigned_tasks(task_name)
    
    # Set dependencies for next robot task
    next_robot_task = dependency_manager.get_current_robot_task()
    if next_robot_task:
        dependency_manager.set_robot_task_dependencies(next_robot_task)
    
    return {
        "status": "success", 
        "task": task_name,
        "next_robot_task": next_robot_task,
        "robot_assigned": dependency_manager.robot_assigned_tasks
    }

@app.get("/check-human-dependency")
def check_human_dependency():
    """Check if current human task dependencies are met"""
    current_task = dependency_manager.get_current_human_task()
    if not current_task:
        return {"allowed": True, "message": "No human tasks remaining"}
    
    print(f"üîç Checking human task dependencies for: '{current_task}'")
    print(f"   Human task dependencies: {dependency_manager.human_task_dependency}")
    print(f"   All finished tasks: {dependency_manager.all_finished_tasks}")
    
    allowed, message = dependency_manager.check_human_dependencies(current_task)
    
    if allowed:
        print(f"‚úÖ All human task dependencies met for '{current_task}'")
    else:
        print(f"‚ùå Human task dependencies not met for '{current_task}': {message}")
    
    return {
        "allowed": allowed,
        "message": message,
        "current_task": current_task,
        "dependencies": dependency_manager.human_task_dependency
    }

@app.get("/check-robot-dependency")
def check_robot_dependency():
    """Check if current robot task dependencies are met"""
    current_task = dependency_manager.get_current_robot_task()
    if not current_task:
        return {"allowed": True, "message": "No robot tasks remaining"}
    
    print(f"üîç Checking robot task dependencies for: '{current_task}'")
    print(f"   Robot task dependencies: {dependency_manager.robot_task_dependency}")
    print(f"   All finished tasks: {dependency_manager.all_finished_tasks}")
    
    allowed, message = dependency_manager.check_robot_dependencies(current_task)
    
    if allowed:
        print(f"‚úÖ All robot task dependencies met for '{current_task}'")
    else:
        print(f"‚ùå Robot task dependencies not met for '{current_task}': {message}")
    
    return {
        "allowed": allowed,
        "message": message,
        "current_task": current_task,
        "dependencies": dependency_manager.robot_task_dependency
    }

@app.get("/get-execution-state")
def get_execution_state():
    """Get current execution state"""
    return {
        "human_assigned_tasks": dependency_manager.human_assigned_tasks,
        "robot_assigned_tasks": dependency_manager.robot_assigned_tasks,
        "all_finished_tasks": dependency_manager.all_finished_tasks,
        "current_human_task": dependency_manager.get_current_human_task(),
        "current_robot_task": dependency_manager.get_current_robot_task(),
        "human_task_dependency": dependency_manager.human_task_dependency,
        "robot_task_dependency": dependency_manager.robot_task_dependency
    }

@app.get("/debug-dependencies")
def debug_dependencies():
    """Debug endpoint to check dependency state"""
    return {
        "dependencies": dependency_manager.dependencies,
        "human_finished": dependency_manager.human_finished_tasks,
        "robot_finished": dependency_manager.robot_finished_tasks,
        "all_finished": dependency_manager.all_finished_tasks
    }

@app.get("/get-task-dependencies")
def get_task_dependencies():
    """Get all task dependencies for frontend grouping"""
    try:
        df = pd.read_excel("tasks.xlsx")
        dependencies = {}
        group_names = {}
        for _, row in df.iterrows():
            task_name = str(row["TaskName"]).strip()
            dep = str(row.get("Dependency", "")).strip()
            group_name = str(row.get("GroupName", "")).strip()
            
            if dep and dep.lower() != "nan":
                dependencies[task_name] = [d.strip() for d in dep.split(",")]
            else:
                dependencies[task_name] = []
            
            if group_name and group_name.lower() != "nan":
                group_names[task_name] = group_name
            else:
                group_names[task_name] = ""
        
        return {"dependencies": dependencies, "group_names": group_names}
    except Exception as e:
        print(f"‚ùå Error loading task dependencies: {e}")
        return {"dependencies": {}, "group_names": {}}


@app.get("/robot/state")
async def robot_state():
    try:
        if ROBOT_CONNECTED:
            return robot_executor.get_state()
        else:
            return {"state": "SIMULATED"}
    except Exception as e:
        return {"state": "UNKNOWN", "error": str(e)}

@app.get("/robot/current_task")
async def get_current_robot_task():
    if ROBOT_CONNECTED:
        return {
            "current_task": robot_executor.current_task,
            "current_task_name": robot_executor.current_task_name,
            "is_running": robot_executor.is_running
        }
    else:
        return {"status": "bypassed", "detail": "Pause simulated"}



@app.get("/robot/current_task")
async def get_current_robot_task():
    if ROBOT_CONNECTED:
        return {"current_task": robot_executor.current_task}
    else:
        return {"status": "bypassed", "detail": "Pause simulated"}


@app.get("/robot/pause")
async def pause_robot():
    if ROBOT_CONNECTED:
        send_dashboard_command("pause")
        return {"status": "success", "detail": "Robot paused"}
    else:
        return {"status": "bypassed", "detail": "Pause simulated"}


@app.get("/robot/resume")
async def resume_robot():
    if ROBOT_CONNECTED:
        send_dashboard_command("play")
        return {"status": "success", "detail": "Robot resumed"}
    else:
        return {"status": "bypassed", "detail": "Resume simulated"}

@app.get("/robot/pause_processing")
async def pause_robot_processing():
    if ROBOT_CONNECTED:
        robot_executor.pause_processing()
        return {"status": "success", "detail": "Robot processing paused"}
    else:
        return {"status": "bypassed", "detail": "Pause processing simulated"}

@app.get("/robot/resume_processing")
async def resume_robot_processing():
    if ROBOT_CONNECTED:
        robot_executor.resume_processing()
        return {"status": "success", "detail": "Robot processing resumed"}
    else:
        return {"status": "bypassed", "detail": "Resume processing simulated"}

@app.get("/robot/execution_state")
async def get_robot_execution_state():
    if ROBOT_CONNECTED:
        return {
            "current_task": robot_executor.current_task,
            "executed_tasks": robot_executor.executed_tasks,
            "queued_tasks": robot_executor.queue
        }
    else:
        return {
            "current_task": "SIMULATED_TASK",
            "executed_tasks": [],
            "queued_tasks": []
        }

@app.post("/robot/reset")
async def reset_robot_state():
    if ROBOT_CONNECTED:
        robot_executor.reset()
        dependency_manager.reset()
        return {"status": "success", "detail": "Robot state and dependency manager reset"}
    else:
        return {"status": "bypassed", "detail": "Robot reset simulated"}

@app.post("/reset-dependencies")
async def reset_dependencies():
    """Reset only the dependency manager (finished task lists)"""
    dependency_manager.reset()
    return {"status": "success", "detail": "Dependency manager reset"}



# Save features
SAVE_PATH = "participant_data.xlsx"

class TaskRecord(BaseModel):
    taskName: str
    allocationValue: int

class SaveRequest(BaseModel):
    participantId: str
    allocationTime: int
    startTime: int
    finishTime: int
    tasks: List[TaskRecord]


@app.get("/participant-count")
def participant_count():
    if not os.path.exists(SAVE_PATH):
        return {"count": 0}
    wb = load_workbook(SAVE_PATH)
    ws = wb.active
    return {"count": ws.max_row - 1}  # Exclude header


@app.get("/previous-allocation")
def get_previous_allocation():
    try:
        if not os.path.exists(SAVE_PATH):
            raise HTTPException(status_code=404, detail="participant_data.xlsx not found")

        df = pd.read_excel(SAVE_PATH)

        if df.empty or len(df.columns) <= 3:
            raise HTTPException(status_code=404, detail="No valid previous allocations found")

        last_row = df.iloc[-1]
        tasks = []

        for col in df.columns[0:]:  # Skip the first 4 columns
            try:
                value = float(last_row[col])
                if not (0 <= value <= 10):  # Ensure it's within slider range
                    value = 5  # Default if invalid
                assigned_to = "Human" if value <= 5 else "Robot"
            except (ValueError, TypeError):
                value = 5
                assigned_to = "Unassigned"

            tasks.append({
                "name": col,
                "assignedTo": assigned_to,
                "sliderValue": value
            })

        # Filter out any invalid values
        for t in tasks:
            if t["sliderValue"] is None or t["sliderValue"] != t["sliderValue"]:  # Check NaN
                t["sliderValue"] = 5

        return tasks

    except HTTPException as e:
        raise e
    except Exception as e:
        print("‚ùå Error reading previous allocation:", e)
        raise HTTPException(status_code=500, detail=f"Internal error: {e}")





@app.post("/save")
def save_to_excel(data: SaveRequest):
    try:
        print(f"üîç Debug: Received save request data: {data}")
        print(f"üîç Debug: Data type: {type(data)}")
        print(f"üîç Debug: Participant ID: {data.participantId}")
        print(f"üîç Debug: Allocation Time: {data.allocationTime}")
        print(f"üîç Debug: Start Time: {data.startTime}")
        print(f"üîç Debug: Finish Time: {data.finishTime}")
        print(f"üîç Debug: Tasks count: {len(data.tasks)}")
        for i, task in enumerate(data.tasks):
            print(f"üîç Debug: Task {i}: {task.taskName} = {task.allocationValue}")
    except Exception as e:
        print(f"‚ùå Error in save endpoint: {e}")
        print(f"üîç Debug: Error type: {type(e)}")
        print(f"üîç Debug: Error details: {str(e)}")
        if "validation" in str(e).lower():
            raise HTTPException(status_code=422, detail=f"Validation error: {str(e)}")
        else:
            raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")
    if os.path.exists(SAVE_PATH):
        wb = load_workbook(SAVE_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Participant ID", "Allocation Time", "Start Time", "Finish Time"])  # Header

    # üîπ Generate next available Participant ID if not provided
    if not data.participantId:
        next_id = f"P{ws.max_row}" if ws.max_row > 1 else "P1"
    else:
        next_id = data.participantId
    
    # Check & add any new task names to header
    existing_headers = [cell.value for cell in ws[1]]
    task_names = [t.taskName for t in data.tasks]
    for name in task_names:
        if name not in existing_headers:
            ws.cell(row=1, column=len(existing_headers)+1, value=name)
            existing_headers.append(name)

    # Build the new row
    row = [""] * len(existing_headers)
    row[0] = next_id
    row[1] = time.strftime("%H:%M:%S", time.localtime(data.allocationTime / 1000))
    row[2] = time.strftime("%H:%M:%S", time.localtime(data.startTime / 1000))
    row[3] = time.strftime("%H:%M:%S", time.localtime(data.finishTime / 1000))

    for task in data.tasks:
        idx = existing_headers.index(task.taskName)
        row[idx] = task.allocationValue

    ws.append(row)
    wb.save(SAVE_PATH)
    return {"status": "success", "participantId": next_id}


# Backend Logging
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook
from datetime import datetime
from threading import Lock
import re

# Global variables
log_lock = Lock()
LOG_FILE = "Log.xlsx"

# Helper function to find the next log filename
def get_next_log_filename():
    files = [f for f in os.listdir('.') if re.match(r'log_\d+\.xlsx$', f)]
    if not files:
        return "log_1.xlsx"
    numbers = [int(re.search(r'log_(\d+)\.xlsx', f).group(1)) for f in files]
    return f"log_{max(numbers)+1}.xlsx"

# Pydantic model for logging event data
class LogEvent(BaseModel):
    participantId: str
    event: str
    details: dict = {}

@app.post("/log-event")
def log_event(data: LogEvent):
    from openpyxl import Workbook, load_workbook
    import shutil

    with log_lock:  # Ensure single write access
        try:
            now = datetime.now()
            time_readable = now.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
            time_ms = int(now.timestamp() * 1000)

            # Create log.xlsx if missing
            if not os.path.exists(LOG_FILE):
                wb = Workbook()
                ws = wb.active
                ws.append(["Participant ID", "Time (ms)", "Time (Readable)", "Event", "Details"])
                wb.save(LOG_FILE)
                wb.close()

            # Open existing log
            try:
                wb = load_workbook(LOG_FILE)
            except Exception as e:
                print(f"‚ùå Log file corrupted: {e}. Starting fresh.")
                wb = Workbook()
                ws = wb.active
                ws.append(["Participant ID", "Time (ms)", "Time (Readable)", "Event", "Details"])
                wb.save(LOG_FILE)
                wb.close()
                wb = load_workbook(LOG_FILE)

            ws = wb.active
            ws.append([data.participantId, time_ms, time_readable, data.event, str(data.details)])
            wb.save(LOG_FILE)

            # Check for rotation
            if ws.max_row > 100:  # Rotate after 100 entries
                new_filename = get_next_log_filename()
                shutil.move(LOG_FILE, new_filename)
                print(f"‚Ñπ Log rotated: {new_filename}")

                # Start new log.xlsx
                wb_new = Workbook()
                ws_new = wb_new.active
                ws_new.append(["Participant ID", "Time (ms)", "Time (Readable)", "Event", "Details"])
                wb_new.save(LOG_FILE)
                wb_new.close()

            wb.close()
            return {"status": "logged"}

        except Exception as e:
            print("‚ùå Error logging event:", e)
            raise HTTPException(status_code=500, detail=f"Failed to log event: {e}")

@app.post("/robot/set_mode")
async def set_robot_mode(mode: str = Body(..., embed=True)):
    if mode.lower() == "orange":
        robot_executor.orange_mode = True
        print("üü† Robot mode set to ORANGE")
    else:
        robot_executor.orange_mode = False
        print("üü° Robot mode set to YELLOW")
    return {"status": "success", "orange_mode": robot_executor.orange_mode}



@app.post("/robot/start")
async def start_robot_tasks(request: Request):
    data = await request.json()
    tasks = data.get("tasks", [])
    robot_tasks = [t for t in tasks if t.get("assignedTo") == "Robot"]

    for task in robot_tasks:
        robot_executor.add_task(task["RobotCode"])

    # Start processing tasks (this will begin dependency checking and execution)
    robot_executor.start_processing()

    return {"status": "success", "queued_tasks": [t["RobotCode"] for t in robot_tasks]}


@app.get("/robot/all_completed")
async def robot_all_completed():
    return {"all_completed": robot_executor.all_tasks_completed}

@app.post("/robot/initialize")
async def initialize_robot(request: Request):
    """
    Initialize robot based on Task Step setting:
    - First: Yellow: Base: -79.98, Shoulder: -91.20, Elbow: -87.97, Wrist 1: -91.69, Wrist2: 84.67, Wrist 3: 815.51
    - Second: Orange: Base: -269.66, Shoulder: -91.23, Elbow: -90.28, Wrist 1: -86.09, Wrist2: 92.69, Wrist 3: 808.82
    """
    try:
        # Get the request data to check taskMode
        data = await request.json()
        task_mode = data.get("taskMode", "")
        
        print(f"üîç Debug - Task Mode from frontend: {task_mode}")
        
        # Determine which position to use based on Task Step
        if task_mode == "Second: Orange":
            # Orange task - use orange position
            joint_positions = [-269.66, -91.23, -90.28, -86.09, 92.69, 808.82]
            task_type = "orange"
        elif task_mode == "First: Yellow":
            # Yellow task - use yellow position
            joint_positions = [-79.98, -91.20, -87.97, -91.69, 84.67, 815.51]
            task_type = "yellow"
        else:
            # Default to yellow position if no task mode specified
            joint_positions = [-79.98, -91.20, -87.97, -91.69, 84.67, 815.51]
            task_type = "default"
        
        # Send joint positions to robot using multiple methods
        from urp_trigger import send_joint_positions_improved, send_joint_positions_urbasic, send_joint_positions_urbasic_simple, activate_gripper, test_robot_connection_detailed, check_robot_state, ROBOT_IP
        
        # Test robot connection first with detailed diagnostics
        if not test_robot_connection_detailed():
            return {"status": "error", "message": f"Robot connection test failed at {ROBOT_IP}. Check console for detailed diagnostics."}
        
        # Check robot state
        print("üîç Checking robot state before sending commands...")
        check_robot_state()
        
        # Try URBasic method first (most reliable)
        print("üîß Attempting movement via URBasic...")
        success = send_joint_positions_urbasic(joint_positions)
        
        # If URBasic fails, try simple URBasic
        if not success:
            print("üîß URBasic failed, trying simple URBasic...")
            success = send_joint_positions_urbasic_simple(joint_positions)
        
        # If both URBasic methods fail, try socket method
        if not success:
            print("üîß URBasic methods failed, trying socket method...")
            success = send_joint_positions_improved(joint_positions)
        
        if not success:
            return {"status": "error", "message": "All movement methods failed. Check robot connection and try again."}
        
        # Activate gripper
        gripper_success = activate_gripper()
        if not gripper_success:
            print("‚ö†Ô∏è Warning: Failed to activate gripper")
        
        print(f"‚úÖ Robot initialized with {task_type} position: {joint_positions}")
        return {
            "status": "success", 
            "task_type": task_type,
            "joint_positions": joint_positions,
            "gripper_activated": gripper_success
        }
        
    except Exception as e:
        print(f"‚ùå Initialize robot failed: {e}")
        return {"status": "error", "message": str(e)}